"""
NSSECU3 - WFP: Application Execution Forensics Parser
BAM / CAM / DAM artifact extraction from Windows registry hives.
Methodology: SANS FOR500 | Exports: JSON + CSV + XLSX (auto)

FIXES & FEATURES:
  - BAM/DAM candidate key lists  : works across all ControlSet variants
  - CAM parsing                  : Capability Access Manager (ConsentStore) per-app usage
  - CAM consent types            : microphone, webcam, location, contacts, etc.
  - Timestamp sanity check       : rejects dates outside 1980-2100
  - col_cell (not c) in autofit  : fixes c() color-helper shadow
  - --exports flag               : avoids shadowing Python built-in format()
  - --include-ueme               : (removed — CAM no longer uses UserAssist)
  - --verbose                    : debug diagnostics with parse counters
  - --admin-check                : privilege check + suggestions
  - Permission denied handler    : actionable fix suggestions
  - _add_detail_sheet()          : DRY Excel sheet creation
  - _print_section()             : DRY console output
"""

import argparse, struct, json, csv, os, sys, string, ctypes
from datetime import datetime, timedelta
from pathlib import Path

try:
    from Registry import Registry
except ImportError:
    sys.exit("pip install python-registry")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("pip install openpyxl")

try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False
    class _Dummy:
        def __getattr__(self, _): return ""
    Fore = Style = _Dummy()

USE_COLOR = True


# ── Helpers ──────────────────────────────────────────────────────────────────

def c(text, color):
    return f"{color}{text}{Style.RESET_ALL}" if (USE_COLOR and HAS_COLOR) else str(text)

def is_admin():
    try:    return ctypes.windll.shell.IsUserAnAdmin()
    except: return False

def suggest_permission_fix(path):
    print(c("\n  ┌─ PERMISSION DENIED:", Fore.YELLOW + Style.BRIGHT))
    if "System32\\config" in str(path):
        print(c("  │ Run as Administrator, then: python app_execution_forensics.py --scan-images", Fore.CYAN))
    if "NTUSER.DAT" in str(path):
        print(c("  │ Mount forensic image and use: --disk DRIVE_LETTER", Fore.CYAN))
    print(c("  └─ Forensic images (--disk) are the preferred method\n", Fore.YELLOW + Style.BRIGHT))

def hex_string(data): return " ".join(f"{b:02X}" for b in data)

def filetime_to_dt(ft):
    if ft == 0:                  return "Never (0x0)"
    if ft == 0x7FFFFFFFFFFFFFFF: return "Never Expires"
    if ft == 0xFFFFFFFFFFFFFFFF: return "Not Set"
    try:
        r = datetime(1601, 1, 1) + __import__("datetime").timedelta(microseconds=ft // 10)
        return f"Invalid (0x{ft:016X})" if not (1980 <= r.year <= 2100) else r.isoformat()
    except: return f"Invalid (0x{ft:016X})"

def read_filetime(data):
    try:    return filetime_to_dt(struct.unpack("<Q", data)[0])
    except: return "Invalid"

def ts(key): return key.timestamp().isoformat() if key.timestamp() else None


# ── Hive opening ──────────────────────────────────────────────────────────────

def _open_hive(hive_path, verbose=False):
    """Open hive and return (reg, tmp_path). tmp_path is always None."""
    return Registry.Registry(hive_path), None

def _close_hive(tmp):
    if tmp:
        try: os.unlink(tmp)
        except: pass


# ── Registry key helpers ──────────────────────────────────────────────────────

def _first_key(reg, candidates):
    for p in candidates:
        try:    return reg.open(p), p
        except Registry.RegistryKeyNotFoundException: pass
    return None, None

BAM_PATHS = [
    "System\\CurrentControlSet\\Services\\bam\\State\\UserSettings",
    "System\\CurrentControlSet\\Services\\bam\\UserSettings",
    "System\\ControlSet001\\Services\\bam\\State\\UserSettings",
    "System\\ControlSet001\\Services\\bam\\UserSettings",
    "System\\ControlSet002\\Services\\bam\\State\\UserSettings",
    "System\\ControlSet002\\Services\\bam\\UserSettings",
    "ControlSet001\\Services\\bam\\State\\UserSettings",
    "ControlSet001\\Services\\bam\\UserSettings",
    "ControlSet002\\Services\\bam\\State\\UserSettings",
    "ControlSet002\\Services\\bam\\UserSettings",
]
DAM_PATHS = [
    "System\\CurrentControlSet\\Services",
    "System\\ControlSet001\\Services",
    "System\\ControlSet002\\Services",
    "ControlSet001\\Services",
    "ControlSet002\\Services",
]

def _bam_path(name):
    if name.startswith("\\??\\"): return name[4:]
    if name.startswith("\\Device\\") or (len(name) >= 3 and name[1:3] == ":\\"): return name
    return None


# ── Artifact parsers ──────────────────────────────────────────────────────────

def parse_bam(reg, verbose=False):
    key, path = _first_key(reg, BAM_PATHS)
    if key is None:
        if verbose: print("    [debug] BAM key not found")
        return []
    if verbose: print(f"    [debug] BAM at: {path}")
    arts, skipped, errors = [], 0, 0
    for sid_key in key.subkeys():
        sid = sid_key.name()
        for v in sid_key.values():
            prog = _bam_path(v.name())
            if not prog: skipped += 1; continue
            try:
                raw = v.value()
                arts.append({"artifact_type": "BAM", "user_sid": sid,
                    "program_path": prog,
                    "last_execution": read_filetime(raw[:8]) if len(raw) >= 8 else "Invalid Data",
                    "raw_hex": hex_string(raw), "registry_path": sid_key.path(),
                    "key_last_modified": ts(sid_key)})
            except: errors += 1
    if verbose: print(f"    [debug] BAM: {len(arts)} kept, {skipped} skipped, {errors} errors")
    return arts


# CAM — Capability Access Manager
# Registry location: SOFTWARE\Microsoft\Windows\CurrentVersion\CapabilityAccessManager\ConsentStore
# Each capability (e.g. microphone, webcam, location) has per-app sub-keys with
# QWORD values LastUsedTimeStart and LastUsedTimeStop (Windows FILETIME).
CAM_ROOT = "Software\\Microsoft\\Windows\\CurrentVersion\\CapabilityAccessManager\\ConsentStore"

def _cam_read_qword(key, value_name):
    """Read a named QWORD value from a registry key; return None on any failure."""
    try:
        v = key.value(value_name)
        raw = v.raw_data()                     # bytes
        if len(raw) >= 8:
            return filetime_to_dt(struct.unpack("<Q", raw[:8])[0])
    except Exception:
        pass
    return None

def parse_cam(reg, verbose=False, source_label="SOFTWARE"):
    """
    Parse Capability Access Manager (CAM) artifacts from the SOFTWARE hive.

    Key path:
        SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\CapabilityAccessManager\\ConsentStore\\<capability>\\<app_or_NonPackaged\\app>

    Each leaf key may carry:
        LastUsedTimeStart  — QWORD FILETIME: most recent access start
        LastUsedTimeStop   — QWORD FILETIME: most recent access end
        Value              — REG_SZ "Allow" or "Deny"

    The include_ueme parameter is kept for API compatibility but unused.
    """
    try:
        consent_root = reg.open(CAM_ROOT)
    except Registry.RegistryKeyNotFoundException:
        if verbose:
            print(f"    [debug] CAM ConsentStore key not found ({CAM_ROOT})")
        return []

    arts, errors = [], 0

    for cap_key in consent_root.subkeys():          # e.g. microphone, webcam, location
        capability = cap_key.name()

        def _process_app_key(app_key, package_type):
            """Extract one CAM artifact from a leaf app key."""
            app_id   = app_key.name()
            consent  = None
            try:    consent = app_key.value("Value").value()
            except Exception: pass

            # Timestamps are stored as raw QWORD (REG_QWORD / REG_BINARY)
            start_ts = _cam_read_qword(app_key, "LastUsedTimeStart")
            stop_ts  = _cam_read_qword(app_key, "LastUsedTimeStop")

            # At minimum we want apps that have a consent entry OR a timestamp
            if consent is None and start_ts is None:
                return None

            return {
                "artifact_type":      "CAM",
                "capability":         capability,
                "app_id":             app_id,
                "package_type":       package_type,      # "Packaged" | "NonPackaged"
                "consent":            consent or "Unknown",
                "last_used_start":    start_ts or "Not Set",
                "last_used_stop":     stop_ts  or "Not Set",
                "source":             source_label,
                "registry_path":      app_key.path(),
                "key_last_modified":  ts(app_key),
            }

        # Direct children are packaged (UWP) apps
        for app_key in cap_key.subkeys():
            if app_key.name() == "NonPackaged":
                # NonPackaged sub-keys are Win32 apps
                for np_key in app_key.subkeys():
                    try:
                        entry = _process_app_key(np_key, "NonPackaged")
                        if entry:
                            arts.append(entry)
                    except Exception:
                        errors += 1
            else:
                try:
                    entry = _process_app_key(app_key, "Packaged")
                    if entry:
                        arts.append(entry)
                except Exception:
                    errors += 1

    if verbose:
        print(f"    [debug] CAM: {len(arts)} entries parsed, {errors} errors")
    return arts


def parse_dam(reg, verbose=False):
    key, path = _first_key(reg, DAM_PATHS)
    if key is None:
        if verbose: print("    [debug] Services key not found")
        return []
    if verbose: print(f"    [debug] DAM at: {path}")
    arts, missing, errors = [], 0, 0
    for svc in key.subkeys():
        try:
            ip = svc.value("ImagePath").value()
            arts.append({"artifact_type": "DAM", "service_name": svc.name(),
                "image_path": ip, "registry_path": svc.path(),
                "key_last_modified": ts(svc),
                "raw_hex": hex_string(ip.encode("utf-8")[:64])})
        except: missing += 1
    if verbose: print(f"    [debug] DAM: {len(arts)} services, {missing} without ImagePath")
    return arts


# ── Hive-level parsers ────────────────────────────────────────────────────────

def _open_safe(path, verbose=False):
    """Open hive; return (reg, tmp). Prints error and returns (None, None) on failure."""
    try:
        return _open_hive(path, verbose)
    except PermissionError:
        print(c(f"[!] Permission denied: {path}", Fore.RED))
        suggest_permission_fix(path)
        return None, None
    except Exception as e:
        print(c(f"[!] Cannot open {path}: {e}", Fore.RED))
        return None, None

def parse_system_hive(path, verbose=False):
    reg, tmp = _open_safe(path, verbose)
    if reg is None: return [], 0, 0
    try:
        b, d = parse_bam(reg, verbose), parse_dam(reg, verbose)
        return b + d, len(b), len(d)
    finally: _close_hive(tmp)

def parse_software_hive(path, verbose=False):
    """Parse the SYSTEM-wide SOFTWARE hive for CAM artifacts."""
    reg, tmp = _open_safe(path, verbose)
    if reg is None: return []
    try:    return parse_cam(reg, verbose, source_label="SOFTWARE")
    finally: _close_hive(tmp)

def parse_ntuser_cam(path, username, verbose=False):
    """Parse a per-user NTUSER.DAT hive for CAM (CapabilityAccessManager\\ConsentStore)."""
    reg, tmp = _open_safe(path, verbose)
    if reg is None: return []
    try:    return parse_cam(reg, verbose, source_label=f"NTUSER({username})")
    finally: _close_hive(tmp)

def parse_single_hive(path, verbose=False):
    reg, tmp = _open_safe(path, verbose)
    if reg is None: return [], 0, 0, 0
    try:
        b  = parse_bam(reg, verbose)
        ca = parse_cam(reg, verbose)
        d  = parse_dam(reg, verbose)
        return b + ca + d, len(b), len(ca), len(d)
    finally: _close_hive(tmp)


# ── Export ────────────────────────────────────────────────────────────────────

def _row(art, trunc=100):
    t = art["artifact_type"]
    base = [t, "", "", "", art.get("last_execution", art.get("last_used_start", "")),
            art.get("registry_path",""), art.get("key_last_modified",""), art.get("raw_hex","")[:trunc]]
    if t == "BAM":
        base[1], base[2] = art["program_path"], art["user_sid"]
    elif t == "CAM":
        base[1] = art.get("capability", "")
        base[2] = art.get("app_id", "")
        base[3] = art.get("source", "")
        base[4] = art.get("last_used_start", "")
    else:
        base[1], base[2] = art["image_path"], art["service_name"]
    return base

def export_csv(arts, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Artifact Type","Program/Capability","App ID","Source Hive",
                    "Last Used Start / Last Execution","Registry Path","Key Modified","Raw Hex"])
        for a in arts: w.writerow(_row(a))

# Excel styles
_H  = {"fill": PatternFill("solid", fgColor="1F4E79"), "font": Font(color="FFFFFF", bold=True, size=11)}
_ST = {
    "BAM": {"fill": PatternFill("solid", fgColor="C6EFCE"), "font": Font(color="006100", bold=True)},
    "CAM": {"fill": PatternFill("solid", fgColor="FFD966"), "font": Font(color="9C5D00", bold=True)},
    "DAM": {"fill": PatternFill("solid", fgColor="B4C6E7"), "font": Font(color="003366", bold=True)},
    "ts":  {"fill": PatternFill("solid", fgColor="FFFFCC")},
    "alt": {"fill": PatternFill("solid", fgColor="F2F2F2")},
}
_BDR = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def _hdr(ws, row, n):
    for col in range(1, n+1):
        cell = ws.cell(row=row, column=col)
        cell.fill, cell.font, cell.border = _H["fill"], _H["font"], _BDR
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def _autofit(ws, mx=60):
    for cols in ws.columns:
        w = max((len(str(col_cell.value or "")) for col_cell in cols), default=10)  # [C] col_cell not c
        ws.column_dimensions[cols[0].column_letter].width = min(w + 3, mx)

def _detail_sheet(wb, title, headers, rows, ts_col=None):
    if not rows: return
    sh = wb.create_sheet(title)
    sh.append(headers)
    _hdr(sh, 1, len(headers))
    for i, row in enumerate(rows, 2):
        sh.append(row)
        for col in range(1, len(headers)+1):
            cell = sh.cell(row=i, column=col)
            cell.border = _BDR
            if ts_col and col == ts_col:
                cell.fill = _ST["ts"]["fill"]
    _autofit(sh)

def export_excel(arts, path):
    wb = Workbook()
    ws = wb.active; ws.title = "All Artifacts"
    hdr = ["Artifact Type","Capability","App ID","Source Hive",
           "Last Used Start / Last Exec","Registry Path","Key Last Modified"]
    ws.append(hdr); _hdr(ws, 1, len(hdr))

    for ri, a in enumerate(arts, 2):
        t = a["artifact_type"]
        if t == "BAM":
            row = [t, a["program_path"], a["user_sid"], "", a["last_execution"], a["registry_path"], a["key_last_modified"]]
        elif t == "CAM":
            row = [t, a.get("capability",""), a.get("app_id",""), a.get("source",""),
                   a.get("last_used_start",""), a["registry_path"], a["key_last_modified"]]
        else:
            row = [t, a["image_path"], a["service_name"], "", "", a["registry_path"], a["key_last_modified"]]
        ws.append(row)
        ws.cell(ri, 1).fill = _ST[t]["fill"]; ws.cell(ri, 1).font = _ST[t]["font"]
        for col in range(1, len(hdr)+1):
            cell = ws.cell(ri, col); cell.border = _BDR
            if col == 5 and row[4]: cell.fill = _ST["ts"]["fill"]
            elif ri % 2 == 0 and cell.fill.start_color.index == "00000000": cell.fill = _ST["alt"]["fill"]
    _autofit(ws, 50)

    _detail_sheet(wb, "BAM Details",
        ["Program Path","User SID","Last Execution","Registry Path","Raw Hex"],
        [[a["program_path"],a["user_sid"],a["last_execution"],a["registry_path"],a["raw_hex"][:60]]
         for a in arts if a["artifact_type"]=="BAM"], ts_col=3)
    _detail_sheet(wb, "CAM Details",
        ["Capability","App ID","Package Type","Source Hive","Consent","Last Used Start","Last Used Stop","Registry Path","Key Modified"],
        [[a.get("capability",""), a.get("app_id",""), a.get("package_type",""),
          a.get("source",""), a.get("consent",""), a.get("last_used_start",""),
          a.get("last_used_stop",""), a["registry_path"], a["key_last_modified"]]
         for a in arts if a["artifact_type"]=="CAM"], ts_col=6)
    _detail_sheet(wb, "DAM Details",
        ["Service Name","Image Path","Registry Path","Key Modified"],
        [[a["service_name"],a["image_path"],a["registry_path"],a["key_last_modified"]]
         for a in arts if a["artifact_type"]=="DAM"])
    wb.save(path)

def _export(arts, stem, args):
    fmts = getattr(args, "export_formats", ["json","csv","xlsx"])
    if not fmts or "all" in fmts: fmts = ["json","csv","xlsx"]
    fmts = list(dict.fromkeys(f.lower() for f in fmts if f))
    for fmt in fmts:
        fp = getattr(args, fmt, None) or f"{stem}.{fmt}"
        if fmt == "json":
            with open(fp, "w", encoding="utf-8") as f: json.dump(arts, f, indent=4, default=str)
        elif fmt == "csv":  export_csv(arts, fp)
        elif fmt == "xlsx": export_excel(arts, fp)
        print(c(f"  [+] {fmt.upper():<5} -> {fp}", Fore.GREEN))


# ── Console output ────────────────────────────────────────────────────────────

def _print_section(title, color, entries, limit, lines_fn, label):
    if not entries: return
    print(c(f"\n  {title:<50}", color + Style.BRIGHT))
    print(c("─"*75, Style.DIM))
    for a in entries[:limit]:
        for line in lines_fn(a): print(line)
        print()
    if len(entries) > limit: print(f"    ... and {len(entries)-limit} more {label} entries\n")

def print_summary(arts, nb, nc, nd):
    print(c(f"\n{'='*75}", Fore.CYAN))
    print(f"  {c('APPLICATION EXECUTION FORENSICS', Fore.CYAN + Style.BRIGHT)}")
    print(c(f"{'='*75}", Fore.CYAN))
    print(f"  BAM: {c(str(nb), Fore.GREEN)}   CAM: {c(str(nc), Fore.GREEN)}   "
          f"DAM: {c(str(nd), Fore.GREEN)}   Total: {c(str(len(arts)), Fore.CYAN + Style.BRIGHT)}\n")

    _print_section("BAM (Background Activity Moderator)", Fore.GREEN,
        [a for a in arts if a["artifact_type"]=="BAM"], 10,
        lambda a: [f"    Program: {c(a['program_path'],Fore.WHITE)}",
                   f"      User:  {c(a['user_sid'],Fore.BLUE)}",
                   f"      Last:  {c(a['last_execution'],Fore.YELLOW)}"], "BAM")
    _print_section("CAM (Capability Access Manager)", Fore.YELLOW,
        [a for a in arts if a["artifact_type"]=="CAM"], 5,
        lambda a: [f"    Capability: {c(a.get('capability',''),Fore.WHITE)}",
                   f"    App:        {c(a.get('app_id',''),Fore.WHITE)}",
                   f"    Type:       {c(a.get('package_type',''),Fore.BLUE)}",
                   f"    Source:     {c(a.get('source',''),Fore.MAGENTA)}",
                   f"    Consent:    {c(a.get('consent',''),Fore.CYAN)}",
                   f"    Last Used:  {c(a.get('last_used_start',''),Fore.YELLOW)}"], "CAM")
    _print_section("DAM (Services/Drivers)", Fore.MAGENTA,
        [a for a in arts if a["artifact_type"]=="DAM"], 10,
        lambda a: [f"    Service: {c(a['service_name'],Fore.BLUE)}",
                   f"      Path:  {c(a['image_path'],Fore.WHITE)}"], "DAM")
    print()


# ── Discovery ─────────────────────────────────────────────────────────────────

def _scan_drives(only=None):
    imgs = []
    for drv in ([only.upper()] if only else list(string.ascii_uppercase)):
        root = f"{drv}:\\"
        if os.path.isdir(os.path.join(root, "Windows", "System32")):
            imgs.append({"root": root, "type": f"drive_{drv}"})
    return imgs

def _find_hives(root):
    cfg = os.path.join(root, "Windows", "System32", "config")
    hives = {"system": None, "ntuser_files": [], "sam": None, "software": None}
    for n, k in [("System","system"),("SAM","sam"),("Software","software")]:
        p = os.path.join(cfg, n)
        if os.path.isfile(p): hives[k] = p
    users = os.path.join(root, "Users")
    if os.path.isdir(users):
        for u in os.listdir(users):
            p = os.path.join(users, u, "NTUSER.DAT")
            if os.path.isfile(p): hives["ntuser_files"].append({"username": u, "path": p})
    return hives

def find_systems(only=None):
    label = f"SCANNING DRIVE {only.upper()}:\\" if only else "SCANNING ALL DRIVES"
    print(c(f"\n  {'='*75}", Fore.CYAN))
    print(f"  {c(label, Fore.CYAN + Style.BRIGHT)}")
    print(c(f"{'='*75}\n", Fore.CYAN))
    found = []
    for img in _scan_drives(only):
        root = img["root"]
        print(f"  {c(img['type'].upper(), Fore.CYAN)}: {c(root, Fore.WHITE)}")
        hives = _find_hives(root)
        if hives["system"]:   print("      \u2713 System hive")
        if hives["sam"]:      print("      \u2713 SAM hive")
        if hives["software"]: print("      \u2713 Software hive")
        if hives["ntuser_files"]:
            print(f"      \u2713 {len(hives['ntuser_files'])} NTUSER.DAT:")
            for u in hives["ntuser_files"]: print(f"        - {c(u['username'], Fore.YELLOW)}")
        print()
        found.append({"info": img, "hives": hives})
    if not found:
        print(c(f"  [!] No Windows system found{' on '+only.upper()+':' if only else ''}.", Fore.RED))
    return found

def _pick_system(systems):
    if len(systems) == 1:
        print(f"  Auto-selecting: {c(systems[0]['info']['root'], Fore.GREEN)}\n")
        return systems[0]
    print(f"\n  {c('Select a system:', Fore.CYAN)}")
    for i, s in enumerate(systems, 1): print(f"    [{i}] {s['info']['root']}")
    print("    [0] Exit")
    while True:
        try:
            ch = int(input("\n  Enter selection: "))
            if ch == 0: return None
            if 1 <= ch <= len(systems): return systems[ch-1]
            print(c("  Invalid. Try again.", Fore.RED))
        except ValueError: print(c("  Invalid. Try again.", Fore.RED))


# ── Analysis pipeline ─────────────────────────────────────────────────────────

def _default_stem(label):
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in label)
    return f"forensics_{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

def _prompt_export(artifacts, default_label, args):
    """
    Interactive terminal export menu shown after the summary.
    Asks the user which format(s) to save and what filename to use.
    Skipped if --out and --exports are both supplied on the command line.
    """
    # If the user pre-supplied both flags, skip the prompt entirely
    cli_out     = getattr(args, "out", None)
    cli_formats = getattr(args, "export_formats", None)
    if cli_out and cli_formats and "all" not in cli_formats:
        _export(artifacts, cli_out, args)
        return

    print(c(f"\n{'─'*75}", Fore.CYAN))
    print(f"  {c('EXPORT OPTIONS', Fore.CYAN + Style.BRIGHT)}")
    print(c(f"{'─'*75}", Fore.CYAN))

    # ── Step 1: choose formats ────────────────────────────────────────────────
    print(f"\n  {c('Which formats would you like to export?', Fore.WHITE)}")
    print(f"    {c('[1]', Fore.CYAN)} JSON")
    print(f"    {c('[2]', Fore.CYAN)} CSV")
    print(f"    {c('[3]', Fore.CYAN)} Excel (.xlsx)")
    print(f"    {c('[4]', Fore.CYAN)} All three  (default)")
    print(f"    {c('[5]', Fore.CYAN)} Pick multiple  (e.g. 1,3 for JSON + Excel)")
    print(f"    {c('[0]', Fore.CYAN)} Skip export")

    FORMAT_MAP = {"1": "json", "2": "csv", "3": "xlsx"}

    while True:
        try:
            choice = input(f"\n  {c('Enter choice:', Fore.WHITE)} ").strip()
        except (EOFError, KeyboardInterrupt):
            print(c("\n  [!] Export skipped.", Fore.YELLOW))
            return

        if choice == "0":
            print(c("  [!] Export skipped.", Fore.YELLOW))
            return
        elif choice in ("", "4"):
            chosen_formats = ["json", "csv", "xlsx"]
            break
        elif choice == "5":
            raw = input(f"  {c('Enter numbers separated by commas (e.g. 1,3):', Fore.WHITE)} ").strip()
            parts = [p.strip() for p in raw.split(",")]
            chosen_formats = [FORMAT_MAP[p] for p in parts if p in FORMAT_MAP]
            if not chosen_formats:
                print(c("  Invalid. Try again.", Fore.RED))
                continue
            break
        elif choice in FORMAT_MAP:
            chosen_formats = [FORMAT_MAP[choice]]
            break
        else:
            print(c("  Invalid. Try again.", Fore.RED))

    # ── Step 2: choose output filename ────────────────────────────────────────
    default_stem = cli_out or _default_stem(default_label)
    print(f"\n  {c('Output filename (no extension):', Fore.WHITE)}")
    print(f"  Press {c('Enter', Fore.CYAN)} to use default: {c(default_stem, Fore.YELLOW)}")

    try:
        name_input = input(f"\n  {c('Filename:', Fore.WHITE)} ").strip()
    except (EOFError, KeyboardInterrupt):
        name_input = ""

    stem = name_input if name_input else default_stem

    # ── Step 3: per-format path overrides (optional) ──────────────────────────
    print(f"\n  {c('Save all files to the current directory?', Fore.WHITE)}")
    print(f"    {c('[1]', Fore.CYAN)} Yes, use current directory  (default)")
    print(f"    {c('[2]', Fore.CYAN)} Specify a folder")

    try:
        dir_choice = input(f"\n  {c('Enter choice:', Fore.WHITE)} ").strip()
    except (EOFError, KeyboardInterrupt):
        dir_choice = "1"

    if dir_choice == "2":
        try:
            folder = input(f"  {c('Folder path:', Fore.WHITE)} ").strip()
        except (EOFError, KeyboardInterrupt):
            folder = ""
        if folder and os.path.isdir(folder):
            stem = os.path.join(folder, os.path.basename(stem))
        elif folder:
            try:
                os.makedirs(folder, exist_ok=True)
                stem = os.path.join(folder, os.path.basename(stem))
                print(c(f"  Created folder: {folder}", Fore.GREEN))
            except Exception as e:
                print(c(f"  [!] Could not create folder ({e}), saving to current directory.", Fore.YELLOW))

    # ── Step 4: export ────────────────────────────────────────────────────────
    print()
    args.export_formats = chosen_formats
    _export(artifacts, stem, args)


def analyze(hives, label, args):
    vb = getattr(args, "verbose", False)
    all_arts, nb, nc, nd = [], 0, 0, 0

    print(c(f"\n{'='*75}", Fore.CYAN))
    print(f"  {c('ANALYZING SELECTED SYSTEM', Fore.CYAN + Style.BRIGHT)}")
    print(c(f"{'='*75}\n", Fore.CYAN))

    if hives.get("system"):
        print(f"  System hive: {c(hives['system'], Fore.WHITE)}")
        try:
            arts, b, d = parse_system_hive(hives["system"], vb)
            all_arts.extend(arts); nb, nd = b, d
            print(f"    \u2713 {c(str(b),Fore.GREEN)} BAM  {c(str(d),Fore.GREEN)} DAM\n")
        except Exception as e: print(c(f"    [!] {e}\n", Fore.YELLOW))

    if hives.get("software"):
        print(f"  Software hive: {c(hives['software'], Fore.WHITE)}")
        try:
            ca = parse_software_hive(hives["software"], vb)
            all_arts.extend(ca); nc += len(ca)
            print(f"    \u2713 {c(str(len(ca)),Fore.GREEN)} CAM (SOFTWARE)\n")
        except Exception as e: print(c(f"    [!] {e}\n", Fore.YELLOW))

    for uh in hives.get("ntuser_files", []):
        print(f"  NTUSER [{uh['username']}]: {c(uh['path'], Fore.WHITE)}")
        try:
            ca = parse_ntuser_cam(uh["path"], uh["username"], vb)
            all_arts.extend(ca); nc += len(ca)
            print(f"    \u2713 {c(str(len(ca)),Fore.GREEN)} CAM (NTUSER)\n")
        except Exception as e: print(c(f"    [!] {e}\n", Fore.YELLOW))

    if not all_arts: print(c("  [!] No artifacts found.", Fore.RED)); return

    print(c(f"\n  {len(all_arts)} artifact(s) found.\n", Fore.GREEN + Style.BRIGHT))
    print_summary(all_arts, nb, nc, nd)
    _prompt_export(all_arts, label, args)
    print()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    global USE_COLOR
    ap = argparse.ArgumentParser(
        description="BAM/CAM/DAM Artifact Parser",
        epilog=("  --disk E               scan drive E only\n"
                "  --scan-images          scan all drives\n"
                "  --scan-images --auto   auto-select first system\n"
                "  path/to/System         parse specific hive\n"),
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("registry", nargs="?", help="Path to a hive file")
    ap.add_argument("--scan-images", action="store_true")
    ap.add_argument("--disk",    metavar="DRIVE")
    ap.add_argument("--auto",    action="store_true")
    ap.add_argument("--out",     metavar="STEM")
    ap.add_argument("--exports", metavar="FORMATS", default="all",
                    help="json,csv,xlsx or all (default)")
    ap.add_argument("--json");  ap.add_argument("--csv"); ap.add_argument("--xlsx")
    ap.add_argument("--no-color",     action="store_true")
    ap.add_argument("--verbose",      action="store_true")
    ap.add_argument("--admin-check",  action="store_true")
    args = ap.parse_args()

    if args.no_color or not HAS_COLOR: USE_COLOR = False

    if args.admin_check:
        ok = is_admin()
        print(c(f"{'✓' if ok else '✗'} {'Running as' if ok else 'NOT'} Administrator",
                Fore.GREEN if ok else Fore.RED))
        if not ok: print(c("  Use --disk for forensic images (no admin needed)", Fore.YELLOW))
        print()

    raw = args.exports.lower().strip()
    args.export_formats = ["json","csv","xlsx"] if raw == "all" else [f.strip() for f in raw.split(",") if f.strip()]
    if args.disk:
        drive = args.disk.strip().rstrip(":\\").upper()
        sys_list = find_systems(only=drive)
        if not sys_list: sys.exit(1)
        analyze(sys_list[0]["hives"], f"drive_{drive}", args); return

    if args.scan_images:
        sys_list = find_systems()
        if not sys_list: sys.exit(1)
        sel = sys_list[0] if args.auto else _pick_system(sys_list)
        if sel is None: sys.exit(0)
        analyze(sel["hives"], sel["info"]["root"].rstrip("\\:"), args); return

    if not args.registry: ap.print_help(); sys.exit(1)
    if not os.path.isfile(args.registry): print(f"[!] Not found: {args.registry}"); sys.exit(1)

    print(c("\n  FORENSICS PARSER", Fore.CYAN + Style.BRIGHT))
    print(f"  Hive: {c(os.path.abspath(args.registry), Fore.WHITE)}\n")
    arts, nb, nc, nd = parse_single_hive(args.registry, args.verbose)
    if not arts: print(c("  [!] No artifacts found.", Fore.RED)); sys.exit(0)
    print(c(f"  {len(arts)} artifact(s) found.\n", Fore.GREEN + Style.BRIGHT))
    print_summary(arts, nb, nc, nd)
    _prompt_export(arts, Path(args.registry).stem, args)
    print()

if __name__ == "__main__":
    main()
